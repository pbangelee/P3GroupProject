#Group 10:
#Kenton Harris, Nathan Saez, Aaron Shumway, Angelee Marshall, Jennica Olsen
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
bold_font = Font(bold=True)  # Bold font

myWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")

currWs = myWorkbook.active

for row in currWs.iter_rows(min_row=2, values_only=True, min_col=1, max_col=3):
    if row[0] not in myWorkbook.sheetnames:
        myWorkbook.create_sheet(row[0])
        myWorkbook.active = myWorkbook[str(row[0])]
        currWs = myWorkbook.active
        currWs["A1"] = "Last Name"
        currWs["B1"] = "First Name"
        currWs["C1"] = "Student ID"
        currWs["D1"] = "Grade"
        myWorkbook.active = myWorkbook["Grades"]
        currWs = myWorkbook.active

for row in currWs.iter_rows(min_row=2, values_only=True, min_col=1, max_col=3):
    studentInfo = row[1].split("_")
    studentInfo.append(row[2])
    myWorkbook.active = myWorkbook[row[0]]
    myWorkbook.active.append(studentInfo)

# Create new worksheets for each class (e.g., a sheet for Algebra, a sheet for Calculus, etc.)

for ws in myWorkbook.worksheets:
    max_row = ws.max_row
    ws.auto_filter.ref = f"A1:D{max_row}"

# Additionally, each sheet should have some simple summary information about each class using functions in columns F (the titles) and G (the data). It should show:
# The highest grade, The lowest grade, The mean grade, The median grade, The number of students in the class

    ws["F1"] = "Summary Statistics"
    ws["G1"] = "Value"
    ws.column_dimensions["F"].width = len(ws["F1"].value) + 5
    ws.column_dimensions["G"].width = len(ws["G1"].value) + 5


    ws["F2"] = "Highest Grade"
    ws["G2"] = f"=MAX(D2:D{max_row})"

    ws["F3"] = "Lowest Grade"
    ws["G3"] = f"=MIN(D2:D{max_row})"
    
    ws["F4"] = "Mean Grade"
    ws["G4"] = f"=AVERAGE(D2:D{max_row})"
    
    ws["F5"] = "Median Grade"
    ws["G5"] = f"=MEDIAN(D2:D{max_row})"

    ws["F6"] = "Number of Students"
    ws["G6"] = f"=COUNT(D2:D{max_row})"

# Some simple formatting (bolding headers) and changing the width of the columns.
# The width of the columns for A,B,C,D,F,G must each be set to the number of characters in the header + 5. 

    column_letters = ["A", "B", "C", "D", "F", "G"]

# Loop through columns A–D formatting width
    for i in range(6): 
        col_letter = column_letters[i]
        header_text = ws[f"{col_letter}1"].value 
        if header_text:
            ws.column_dimensions[col_letter].width = len(header_text) + 5

# Bolding the headers
    for cell in ws[1]:
        cell.font = bold_font

myWorkbook.remove(myWorkbook["Grades"])

myWorkbook.save(filename="formatted_grades.xlsx")
myWorkbook.close()