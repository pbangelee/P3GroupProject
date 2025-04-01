#Group 10:
#Kenton Harris, Nathan Saez, Aaron Shumway, Angelee Marshall, Jennica Olsen
# Imports excel file and takes data, formats it, and creates new excel file that is more organized
import openpyxl
from openpyxl.styles import Font
bold_font = Font(bold=True)  # Creates Bold font object

# Imports poorly organized file and sets it equal to myWorkbook variable
myWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")

# sets currws equal to active sheet
currWs = myWorkbook.active

# Iterates through all the rows and creates new sheet for the class if the sheet hasn't been made
# Sets header values as each sheet is made
for row in currWs.iter_rows(min_row=2, values_only=True, min_col=1, max_col=3):
    if row[0] not in myWorkbook.sheetnames:
        myWorkbook.create_sheet(row[0])
        myWorkbook.active = myWorkbook[str(row[0])]
        currWs = myWorkbook.active
        currWs["A1"] = "Last Name"
        currWs["B1"] = "First Name"
        currWs["C1"] = "Student ID"
        currWs["D1"] = "Grade"
        myWorkbook.active = myWorkbook["Grades"]
        currWs = myWorkbook.active

# Iterates through the rows and splits the data into a list variable
for row in currWs.iter_rows(min_row=2, values_only=True, min_col=1, max_col=3):
    studentInfo = row[1].split("_")
    studentInfo.append(row[2]) # Adds grades to end of list
    myWorkbook.active = myWorkbook[row[0]]
    myWorkbook.active.append(studentInfo) # Adds list data to next available row in sheet

# Goes through each sheet and adds filter the the disgnated row and assigns excel functions
for ws in myWorkbook.worksheets:
    max_row = ws.max_row
    ws.auto_filter.ref = f"A1:D{max_row}" # Applies filter

    ws["F1"] = "Summary Statistics"
    ws["G1"] = "Value"
    ws.column_dimensions["F"].width = len(ws["F1"].value) + 5 # Changes the formatting/width of column
    ws.column_dimensions["G"].width = len(ws["G1"].value) + 5


    ws["F2"] = "Highest Grade"
    ws["G2"] = f"=MAX(D2:D{max_row})" # Max function

    ws["F3"] = "Lowest Grade"
    ws["G3"] = f"=MIN(D2:D{max_row})" # Min function
    
    ws["F4"] = "Mean Grade"
    ws["G4"] = f"=AVERAGE(D2:D{max_row})" # Average function
    
    ws["F5"] = "Median Grade"
    ws["G5"] = f"=MEDIAN(D2:D{max_row})" # Median function

    ws["F6"] = "Number of Students"
    ws["G6"] = f"=COUNT(D2:D{max_row})" # Count function

    # Makes list of column letters
    column_letters = ["A", "B", "C", "D"]

# Loop through columns A–D formatting width
    for i in range(4): 
        col_letter = column_letters[i] # sets column letter
        header_text = ws[f"{col_letter}1"].value # sets the headers value using column letter
        if header_text:
            ws.column_dimensions[col_letter].width = len(header_text) + 5 # Adjusts column width

# Bolding the headers
    for cell in ws[1]:
        cell.font = bold_font

# Removes initial sheet
myWorkbook.remove(myWorkbook["Grades"])

# Save and close file and names it
myWorkbook.save(filename="formatted_grades.xlsx")
myWorkbook.close()